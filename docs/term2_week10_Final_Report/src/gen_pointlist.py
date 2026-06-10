#!/usr/bin/env python3
"""
Generate the Appendix B point-list tables for the final report from the Excel
workbook. Reads PLC/src/Point_List.xlsx and writes appendix_pointlist.tex, which
the report \input{}s. Every sheet becomes a landscape, full-grid table using the
spreadsheet's columns (Point Description, Origin Address, DO/DI/AO/AI/Pwr,
Destination Address, Destination Description, Notes).

Usage (run from docs/term2_week10_Final_Report/src/):
    python3 gen_pointlist.py            # uses default paths
    python3 gen_pointlist.py ../../../PLC/src/Point\\ List.xlsx appendix_pointlist.tex

Requires: openpyxl  (pip install openpyxl)
"""
import sys
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "../../../PLC/src/Point_List.xlsx"
OUT  = sys.argv[2] if len(sys.argv) > 2 else "appendix_pointlist.tex"

# Unicode -> LaTeX replacements for spreadsheet text
UNI = {
    '—': '--', '–': '--', '‒': '--', '″': "''", '′': "'",
    '→': r'$\rightarrow$', '←': r'$\leftarrow$',
    '≥': r'$\geq$', '≤': r'$\leq$',
    'Ω': r'$\Omega$', '°': r'\textdegree{}', 'µ': r'$\mu$',
    '×': r'$\times$', '…': '...', '•': r'\textbullet{}',
    '“': "``", '”': "''", '‘': "`", '’': "'",
    ' ': ' ', '±': r'$\pm$',
}

def esc(v):
    if v is None:
        return ''
    s = str(v)
    if s.strip() == '':
        return ''
    out = []
    for ch in s:
        # Unicode -> LaTeX substitutions go in raw (no further escaping)
        if ch in UNI:
            out.append(UNI[ch])
        elif ch in '&%$#_{}':
            out.append('\\' + ch)
        elif ch == '~': out.append(r'\textasciitilde{}')
        elif ch == '^': out.append(r'\textasciicircum{}')
        elif ch == '\\': out.append(r'\textbackslash{}')
        elif ch == '>': out.append(r'\textgreater{}')
        elif ch == '<': out.append(r'\textless{}')
        else: out.append(ch)
    return ''.join(out)

def cell(ws, r, c):
    return ws.cell(row=r, column=c).value

def num(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)

COLSPEC = r'|p{3.0cm}|p{1.25cm}|c|c|c|c|c|p{2.5cm}|p{4.0cm}|X|'
HEADERS = ['Point Description', 'Origin', 'DO', 'DI', 'AO', 'AI', 'Pwr',
           'Destination Address', 'Destination Description', 'Notes']

def header_block():
    h = ' & '.join(r'\textbf{%s}' % x for x in HEADERS)
    return ('\\hline\n\\rowcolor{PSUgreen!12}' + h + ' \\\\\n\\hline\n\\endfirsthead\n'
            '\\hline\n\\rowcolor{PSUgreen!12}' + h + ' \\\\\n\\hline\n\\endhead\n')

def find_header_row(ws):
    for r in range(1, min(ws.max_row, 8) + 1):
        for c in range(1, ws.max_column + 1):
            v = cell(ws, r, c)
            if v and 'Point Description' in str(v):
                return r
    return None

def _is_total_row(vals):
    """Detect a stale Total/Sum row in the spreadsheet so we can drop it."""
    head = (str(vals[0] or '') + ' ' + str(vals[1] or '')).strip().lower()
    return head.startswith('total') or head.startswith('sum') or 'total points' in head

def _is_excluded_from_total(vals):
    """Spare / unused / n-a rows: render them, but skip from the recomputed total."""
    desc = str(vals[0] or '').strip().lower()
    return ('spare' in desc) or ('unused' in desc) or desc in ('n/a', 'na', '-')

def _as_int(v):
    """Best-effort int coercion for count cells. Blank / non-numeric -> 0."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0

def emit_standard(ws, hr, ncols):
    # 11-col sheets carry an extra "virtual point" column (skip col 8)
    cmap = [1, 2, 3, 4, 5, 6, 7, 9, 10, 11] if ncols >= 11 else [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    rows = []
    totals = [0, 0, 0, 0, 0]   # DO, DI, AO, AI, Pwr
    for r in range(hr + 1, ws.max_row + 1):
        vals = [cell(ws, r, c) for c in cmap]
        if all(v is None or str(v).strip() == '' for v in vals):
            continue
        if _is_total_row(vals):
            continue  # drop stale total row; we recompute below
        cells = ([esc(vals[0]), esc(vals[1])]
                 + [num(vals[i]) for i in range(2, 7)]
                 + [esc(vals[7]), esc(vals[8]), esc(vals[9])])
        rows.append(' & '.join(cells) + r' \\ \hline')
        if not _is_excluded_from_total(vals):
            for i, t in enumerate(totals):
                totals[i] = t + _as_int(vals[2 + i])
    # Recomputed total row -- bold, light-green fill, sums in the count columns
    total_cells = ([r'\textbf{Total Points}', '']
                   + [r'\textbf{%d}' % t for t in totals]
                   + ['', '', ''])
    rows.append(r'\rowcolor{PSUgreen!12}' + ' & '.join(total_cells) + r' \\ \hline')
    return rows

def emit_generic(ws):
    """Render generic (non-pointlist) sheets. Splits into two blocks at a 'REV'
    header so revision tables get their own page on the Index sheet."""
    maxc = min(ws.max_column, 6)
    blocks = [[]]
    for r in range(1, ws.max_row + 1):
        vals = [cell(ws, r, c) for c in range(1, maxc + 1)]
        if all(v is None or str(v).strip() == '' for v in vals):
            continue
        # If this row looks like a fresh header (REV ... DATE ... DESCRIPTION ...),
        # start a new block so the revision history breaks to its own page.
        # 'REV' lives in column A or B depending on the sheet layout.
        row_strs = [str(v or '').strip().upper() for v in vals[:2]]
        if blocks[-1] and 'REV' in row_strs:
            blocks.append([])
        blocks[-1].append(' & '.join(esc(v) for v in vals) + r' \\ \hline')
    return blocks, maxc

def emit_index(wb):
    """Auto-generate the Index block from the workbook's actual sheet list,
    so the hand-entered PAGE NO. / NO. OF SHEETS columns in the xlsx can't
    drift out of sync. Numbering is sequential and skips the Index sheet itself."""
    rows = [r'\rowcolor{PSUgreen!12}\textbf{Sheet \#} & \textbf{Sheet Name} \\ \hline']
    n = 0
    for nm in wb.sheetnames:
        if nm.strip().lower() == 'index':
            continue
        n += 1
        rows.append('%d & %s \\\\ \\hline' % (n, esc(nm)))
    return rows

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    out = ['% Auto-generated by gen_pointlist.py from PLC/Point_List.xlsx -- do not edit by hand.',
           r'\begingroup', r'\footnotesize', r'\setlength{\tabcolsep}{3pt}',
           r'\renewcommand{\arraystretch}{1.15}', r'\begin{landscape}']
    for i, nm in enumerate(wb.sheetnames):
        ws = wb[nm]
        if i > 0:
            out.append(r'\clearpage')
        out.append('')
        out.append(r'\noindent\textbf{\textcolor{PSUgreen}{\Large Sheet: %s}}\par\medskip' % esc(nm))
        hr = find_header_row(ws)
        if hr is None:
            blocks, maxc = emit_generic(ws)
            if nm.strip().lower() == 'index':
                # Override the stale PAGE NO. / NO. OF SHEETS index with an
                # auto-built one; keep the workbook's Revision-History blocks.
                index_rows = emit_index(wb)
                out += [r'\begin{xltabular}{\linewidth}{|p{2.5cm}|X|}', r'\hline'] + index_rows + [r'\end{xltabular}']
                blocks = blocks[1:]
                maxc = max(maxc, 6)
            spec = '|' + '|'.join(['p{4cm}'] * (maxc - 1) + ['X']) + '|' if maxc > 1 else '|X|'
            for bi, block_rows in enumerate(blocks):
                if bi > 0 or (nm.strip().lower() == 'index' and bi == 0):
                    out.append(r'\clearpage')
                    out.append(r'\noindent\textbf{\textcolor{PSUgreen}{\Large Sheet: %s --- Revision History}}\par\medskip' % esc(nm))
                out += [r'\begin{xltabular}{\linewidth}{%s}' % spec, r'\hline'] + block_rows + [r'\end{xltabular}']
        else:
            rows = emit_standard(ws, hr, ws.max_column)
            out += [r'\begin{xltabular}{\linewidth}{%s}' % COLSPEC, header_block()] + rows + [r'\end{xltabular}']
    out += [r'\end{landscape}', r'\endgroup']
    with open(OUT, 'w') as f:
        f.write('\n'.join(out) + '\n')
    print('Wrote %s (%d sheets)' % (OUT, len(wb.sheetnames)))

if __name__ == '__main__':
    main()
